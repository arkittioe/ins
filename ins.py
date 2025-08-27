import os
import sys
import json
import openpyxl

# ---------- مسیر پایه برای فایل‌ها ----------
if getattr(sys, 'frozen', False):
    base_path = os.path.dirname(sys.executable)  # مسیر exe
else:
    base_path = os.path.dirname(os.path.abspath(__file__))  # مسیر فایل پایتون

# ---------- خواندن تنظیمات از JSON ----------
config_file = os.path.join(base_path, "config.json")

if not os.path.exists(config_file):
    default_config = {
        "base_file": "INS(xxx).xlsx",
        "ref_file": "PaintCalculateFinal-01.xlsx",
        "row_start": 6,
        "max_elements": 40,
        "columns": {
            "lookup": {"desc": "D", "sharh": "F", "tol": "H", "vazn": "L"},
            "output": {"desc": "D", "sharh": "G", "tol": "I", "tedad": "J", "vazn": "K"}
        },
        "output": {"folder": ".", "pattern": "INS({num}).xlsx"}
    }
    with open(config_file, "w", encoding="utf-8") as f:
        json.dump(default_config, f, indent=4, ensure_ascii=False)
    print(f"{config_file} created. Edit it to customize columns and paths.")
    raise SystemExit

with open(config_file, "r", encoding="utf-8") as f:
    config = json.load(f)

base_file = os.path.join(base_path, config["base_file"])
ref_file = os.path.join(base_path, config["ref_file"])
row_start = config.get("row_start", 6)
max_elements = config.get("max_elements", 40)
ref_cols = config["columns"]["lookup"]
out_cols = config["columns"]["output"]

# ---------- بررسی وجود فایل‌ها ----------
if not os.path.exists(base_file):
    raise SystemExit(f"File not found: {base_file}")
if not os.path.exists(ref_file):
    raise SystemExit(f"Reference file not found: {ref_file}")

# ---------- انتخاب شماره INS ----------
num_out = input("Number for INS: ").strip()

# ---------- مسیر و نام خروجی ----------
output_pattern = config.get("output", {}).get("pattern", "INS({num}).xlsx")
output_folder = os.path.join(base_path, config.get("output", {}).get("folder", "."))
os.makedirs(output_folder, exist_ok=True)  # ایجاد پوشه خروجی در صورت عدم وجود
output_file = os.path.join(output_folder, output_pattern.format(num=num_out))

# ---------- بارگذاری فایل‌ها ----------
wb = openpyxl.load_workbook(base_file)
ws = wb.active

ref_wb = openpyxl.load_workbook(ref_file, data_only=True)
ref_ws = ref_wb.active

ws.title = f"نصب {num_out}"
ws["E2"].value = f"شماره صورتمجلس\nINS-{num_out}"

# ---------- دریافت چندین شرح از کاربر ----------
print("Paste your element descriptions (one per line), then enter an empty line to finish:")
descriptions = []
while True:
    line = input().strip()
    if not line:
        break
    descriptions.append(line)

elements_entered = 0

# ---------- وارد کردن داده‌ها در اکسل ----------
for i, desc in enumerate(descriptions, start=1):
    row = row_start + (i - 1)

    sharh = ""
    tol = ""
    vazn = ""

    # جستجو در فایل مرجع
    for r in range(2, ref_ws.max_row + 1):
        if str(ref_ws[f"{ref_cols['desc']}{r}"].value).strip() == desc.strip():
            sharh = ref_ws[f"{ref_cols['sharh']}{r}"].value
            sharh = f"UC{sharh}" if sharh else ""
            tol_val = ref_ws[f"{ref_cols['tol']}{r}"].value
            tol = float(tol_val) / 1000 if tol_val else ""
            vazn = ref_ws[f"{ref_cols['vazn']}{r}"].value
            break

    tedad = input(f"Enter quantity for '{desc}' (row {row}): ")

    # نوشتن در فایل خروجی
    ws[f"{out_cols['desc']}{row}"].value = desc
    ws[f"{out_cols['sharh']}{row}"].value = sharh
    ws[f"{out_cols['tol']}{row}"].value = tol
    ws[f"{out_cols['tedad']}{row}"].value = tedad
    ws[f"{out_cols['vazn']}{row}"].value = vazn

# ---------- ذخیره فایل خروجی ----------
wb.save(output_file)
print(f"Output file created: {output_file}")

# pyinstaller --onefile ins.py