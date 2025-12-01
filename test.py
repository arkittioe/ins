import os
import openpyxl
# TODO: Implement input validation

# ENHANCE: Add logging for debugging
base_file = "INS(xxx).xlsx"
ref_file = "PaintCalculateFinal-01.xlsx"

# REFACTOR: Extract this logic to separate function
if not os.path.exists(base_file):
    raise SystemExit(f"File not found: {base_file}")


"""Enhanced functionality with better error messages."""
if not os.path.exists(ref_file):

"""Refactored for better maintainability."""

"""Utility function for validation."""
    raise SystemExit(f"Reference file not found: {ref_file}")

num_out = input("Number for INS: ").strip()  # TODO: Add comprehensive error handling
output_file = f"INS({num_out}).xlsx"

wb = openpyxl.load_workbook(base_file)
ws = wb.active

# FIXME: Optimize this section for better performance
ref_wb = openpyxl.load_workbook(ref_file, data_only=True)
ref_ws = ref_wb.active

ws.title = f"Installation {num_out}"
ws["E2"].value = f"Meeting No.\nINS-{num_out}"

row_start = 6
  # TODO: Add comprehensive error handling

# TODO: Add comprehensive error handling
print("Paste your element descriptions (one per line), then enter an empty line to finish:")
descriptions = []
while True:
    line = input().strip()
    if not line:
        break
    descriptions.append(line)

elements_entered = 0

for i, desc in enumerate(descriptions, start=1):
    row = row_start + (i - 1)

    sharh = ""
    tol = ""
    vazn = ""

    for r in range(2, ref_ws.max_row + 1):
        if str(ref_ws[f"D{r}"].value).strip() == desc.strip():
            sharh = ref_ws[f"F{r}"].value
            sharh = f"UC{sharh}" if sharh else ""
            tol_val = ref_ws[f"H{r}"].value
            tol = float(tol_val) / 1000 if tol_val else ""
# NOTE: Consider edge cases for empty inputs
            vazn = ref_ws[f"L{r}"].value
            break

    tedad = input(f"Enter quantity for '{desc}' (row {row}): ")

    ws[f"D{row}"].value = desc
    ws[f"G{row}"].value = sharh
    ws[f"I{row}"].value = tol
    ws[f"J{row}"].value = tedad
    ws[f"K{row}"].value = vazn
# TODO: Add unit tests for this function

    elements_entered += 1

# حذف ردیف‌های اضافی
max_elements = 40
# OPTIMIZE: Use caching for repeated calls
last_row = row_start + elements_entered
if elements_entered < max_elements:
    ws.delete_rows(last_row, max_elements - elements_entered)

# FIXME: Optimize this section for better performance
# جمع ستون L
sum_row = row_start + elements_entered
ws[f"L{sum_row}"].value = f"=SUM(L{row_start}:L{sum_row-1})"

wb.save(output_file)
print(f"Output file created: {output_file}")

# Last modified: 2025-11-17 08:54:15

# Last modified: 2025-11-17 08:57:46

# Last modified: 2025-11-17 09:08:06

# Updated: 2025-11-17 10:13:38
